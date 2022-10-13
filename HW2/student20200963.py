import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']
total = []

row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total.append(sum_v) 
	row_id += 1

total.sort(reverse=True)

grade = []
all = len(total)
num = all * 0.3
g = 'A0'
i = 0
while i < all:
	if int(num - total.count(total[i])) >= 0:
		row_id = 2
		count = 0
		for row in ws:
			if total[i] == ws.cell(row_id, column = 7).value:
				ws.cell(row_id, column = 8).value = g
				grade.append(g)
				num -= 1
				count += 1
			row_id += 1
		i += count
	elif g == 'A0':
		num += all * 0.4
		g = 'B0'
	else:
		break

while i < all:
	row_id = 2
	count = 0
	for row in ws:
		if total[i] == ws.cell(row_id, column = 7).value:
			ws.cell(row_id, column = 8).value = 'C'
			grade.append('C0')
			count += 1
		row_id += 1
	i += count

num = grade.count('A0') * 0.5
g = 'A+'
i = 0
while i < all:
	if int(num - total.count(total[i])) >= 0:
		row_id = 2
		count = 0
		for row in ws:
			if total[i] == ws.cell(row_id, column = 7).value:
				ws.cell(row_id, column = 8).value = g
				num -= 1
				count += 1
			row_id += 1
		i += count
	elif g == 'A+':
		i += grade.count('A0')
		num = grade.count('B0') * 0.5
		g = 'B+'
	elif g == 'B+':
		i += grade.count('B0')
		num = grade.count('C0') * 0.5
		g = 'C+'
	else:
		break

wb.save("student.xlsx")
